"""
src/data/loader.py
==================
Section-based Excel parser for data.xlsx.

data.xlsx uses a single sheet ("Part 1") with three logically separate
sections stacked vertically — NOT three separate Excel sheets.  Each section
begins with a named header row in column B:

  Row  1 – "Electrical Components"  (data rows 2-11, total row 12)
  Row 15 – "Mechanical Components"  (data rows 16-30, total row 31)
  Row 33 – "Hybrid Components"      (data rows 34-49, total row 50)

A battery/tank lookup table occupies columns L-R, rows 3-14 (header row 3,
data rows 4-14).

An "Energy" sheet provides per-subsystem shaft power, drive type, drivetrain
efficiency, and selected turbine size for all three systems.

Non-numeric values ("indefinite", "~15 tons", "$ 2500 per ton", etc.) are
stored as-is.  No coercion is performed here; downstream modules handle it.
"""

from pathlib import Path

import openpyxl
import pandas as pd

from src.config import DATA_FILE

# ──────────────────────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────────────────────

# Exact header strings as they appear in column B of Part 1.
# Key is the header text; value is the canonical dict key we expose.
SECTION_HEADERS: dict[str, str] = {
    "Electrical Components": "electrical",
    "Mechanical Components": "mechanical",
    "Hybrid Components": "hybrid",
}

# Ordered column names for the equipment rows (columns B-E, positions 2-5).
EQUIPMENT_COLUMNS = [
    "name",
    "quantity",
    "cost_usd",
    "lifespan_years",
]

# Column names for the battery/tank lookup table (columns J-P, positions 10-16).
BATTERY_COLUMNS = [
    "battery_fraction",
    "tank_fraction",
    "battery_kwh",
    "tank_gal",
    "battery_cost",
    "tank_cost",
    "total_cost",
]

# Column names for the Part 2 TDS and depth lookup tables.
TDS_LOOKUP_COLUMNS = ["tds_ppm", "ro_energy_kw"]
DEPTH_LOOKUP_COLUMNS = ["depth_m", "pump_energy_kw"]


# ──────────────────────────────────────────────────────────────────────────────
# Private helpers
# ──────────────────────────────────────────────────────────────────────────────

def _parse_section(ws, header_row: int, stop_rows: set, cost_col: int = 4) -> list[dict]:
    """
    Parse equipment rows starting immediately after *header_row*.

    Iteration stops when:
      - The current row number is in *stop_rows* (start of next section), OR
      - Column B value is ``None`` (blank cell, including merged-cell spillover).

    Rows where column B value is ``"Total"`` are silently skipped (they are
    section summary rows, not equipment data).

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
    header_row : int
        1-based row number of the section header ("Electrical Components", etc.)
    stop_rows : set[int]
        Row numbers that mark the start of the *next* section (or end boundary).
    cost_col : int
        1-based column index for the cost_usd value.
        Electrical section uses col 5 (E = total cost).
        Mechanical and hybrid sections use col 4 (D = cost).
        Lifespan is always the column immediately after cost_col.

    Returns
    -------
    list[dict]
        One dict per equipment row, keys matching EQUIPMENT_COLUMNS.
    """
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        if r in stop_rows:
            break
        name = ws.cell(r, 2).value
        # None means blank (or merged-cell note row) — stop scanning.
        if name is None:
            break
        # Skip section-total summary rows.
        if name == "Total":
            continue
        rows.append({
            "name":           name,
            "quantity":       ws.cell(r, 3).value,
            "cost_usd":       ws.cell(r, cost_col).value,
            "lifespan_years": ws.cell(r, cost_col + 1).value,
        })
    return rows


def _parse_battery_lookup(ws) -> pd.DataFrame:
    """
    Parse the battery/tank lookup table from the fixed range L3:R14.

    Header labels are at row 3 (cols L-R); data occupies rows 4-14
    (11 rows: battery_fraction 0.0 to 1.0 in 0.1 steps).

    Returns
    -------
    pd.DataFrame  with columns matching BATTERY_COLUMNS
    """
    rows = []
    for r in range(4, 15):   # rows 4–14 inclusive
        rows.append({
            "battery_fraction": ws.cell(r, 12).value,
            "tank_fraction":    ws.cell(r, 13).value,
            "battery_kwh":      ws.cell(r, 14).value,
            "tank_gal":         ws.cell(r, 15).value,
            "battery_cost":     ws.cell(r, 16).value,
            "tank_cost":        ws.cell(r, 17).value,
            "total_cost":       ws.cell(r, 18).value,
        })
    return pd.DataFrame(rows, columns=BATTERY_COLUMNS)


def _parse_part2_lookups(wb) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Parse both lookup tables from the 'Part 2' sheet.

    Layout (confirmed from data.xlsx):
      Row 1: headers — col A="TDS", col B="kW required (RO Desalination)",
                       col D="Depth", col E="kW required (pump energy)"
      Rows 2-21: 20 data rows for each table (values 0-1900 in 100-unit steps)

    Returns
    -------
    tuple[pd.DataFrame, pd.DataFrame]
        (tds_df, depth_df)
        tds_df columns:   ["tds_ppm", "ro_energy_kw"]
        depth_df columns: ["depth_m", "pump_energy_kw"]
    """
    if "Part 2" not in wb.sheetnames:
        raise ValueError(
            f"Expected sheet 'Part 2' not found. "
            f"Available sheets: {wb.sheetnames}"
        )
    ws2 = wb["Part 2"]

    tds_rows = []
    depth_rows = []
    for r in range(2, 22):   # rows 2-21 inclusive (20 data rows)
        tds_rows.append({
            "tds_ppm":      ws2.cell(r, 1).value,   # col A
            "ro_energy_kw": ws2.cell(r, 2).value,   # col B
        })
        depth_rows.append({
            "depth_m":        ws2.cell(r, 4).value, # col D
            "pump_energy_kw": ws2.cell(r, 5).value, # col E
        })

    tds_df   = pd.DataFrame(tds_rows,   columns=TDS_LOOKUP_COLUMNS)
    depth_df = pd.DataFrame(depth_rows, columns=DEPTH_LOOKUP_COLUMNS)

    print(f"  [loader] TDS lookup:   {len(tds_df)} rows parsed")
    print(f"  [loader] Depth lookup: {len(depth_df)} rows parsed")
    return tds_df, depth_df


def _parse_energy_sheet(wb) -> dict | None:
    """
    Parse the Energy sheet into a structured dict grouped by system.

    The Energy sheet has the following row structure:
      - System header rows (e.g., "Mechanical System") in column A
      - Subsystem rows with shaft power, drive type, efficiency, turbine input
      - Summary rows: "Total Shaft Power", "Total at Turbine Shaft" (or variant),
        "Design Power (+10% margin)", "Selected Turbine (kW)"

    Column mapping (1-based):
      A (1) = system/subsystem label
      B (2) = shaft_power_kw
      C (3) = drive_type
      D (4) = drivetrain_efficiency
      E (5) = turbine_input_kw
      F (6) = notes

    Returns
    -------
    dict with keys "mechanical", "electrical", "hybrid", each containing:
        "subsystems"          – list of dicts with subsystem data
        "total_shaft_power"   – float
        "total_turbine_input" – float  (total at turbine shaft before margin)
        "selected_turbine_kw" – float
    Returns None if the Energy sheet is absent (power data falls back to
    SUBSYSTEM_POWER constants in config.py).
    """
    if "Energy" not in wb.sheetnames:
        print("  [loader] Energy sheet not found — using config constants for power data")
        return None

    ws = wb["Energy"]

    # Canonical system header strings → internal keys
    SYSTEM_HEADER_MAP = {
        "Mechanical System": "mechanical",
        "Electrical System": "electrical",
        "Hybrid System":     "hybrid",
    }

    # Summary row labels we capture (partial match using startswith)
    SUMMARY_TOTAL_SHAFT   = "Total Shaft Power"
    SUMMARY_TOTAL_TURBINE = "Total at Turbine"   # covers both variants
    SUMMARY_SELECTED      = "Selected Turbine (kW)"

    result: dict = {}
    current_system: str | None = None
    current_subsystems: list = []
    current_total_shaft: float = 0.0
    current_total_turbine: float = 0.0
    current_selected: float = 0.0

    def _flush(system_key: str) -> None:
        result[system_key] = {
            "subsystems":          list(current_subsystems),
            "total_shaft_power":   current_total_shaft,
            "total_turbine_input": current_total_turbine,
            "selected_turbine_kw": current_selected,
        }

    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue

        label_str = str(label).strip()

        # ── System header row ──────────────────────────────────────────────
        if label_str in SYSTEM_HEADER_MAP:
            if current_system is not None:
                _flush(current_system)
            current_system = SYSTEM_HEADER_MAP[label_str]
            current_subsystems = []
            current_total_shaft = 0.0
            current_total_turbine = 0.0
            current_selected = 0.0
            continue

        if current_system is None:
            continue

        # ── Summary rows ───────────────────────────────────────────────────
        if label_str.startswith(SUMMARY_TOTAL_SHAFT):
            val = ws.cell(r, 2).value
            if val is not None:
                current_total_shaft = float(val)
            continue

        if label_str.startswith(SUMMARY_TOTAL_TURBINE):
            val = ws.cell(r, 5).value
            if val is not None:
                current_total_turbine = float(val)
            continue

        if label_str.startswith(SUMMARY_SELECTED):
            val = ws.cell(r, 5).value
            if val is not None:
                current_selected = float(val)
            continue

        # Skip other summary rows (Design Power, etc.)
        if label_str.startswith("Design Power") or label_str.startswith("Total Electrical"):
            continue

        # ── Subsystem data row ─────────────────────────────────────────────
        shaft_val   = ws.cell(r, 2).value
        turbine_val = ws.cell(r, 5).value
        current_subsystems.append({
            "name":                  label_str,
            "shaft_power_kw":        float(shaft_val)   if shaft_val   is not None else None,
            "drive_type":            ws.cell(r, 3).value,
            "drivetrain_efficiency": ws.cell(r, 4).value,
            "turbine_input_kw":      float(turbine_val) if turbine_val is not None else None,
            "notes":                 ws.cell(r, 6).value,
        })

    # Flush the last system
    if current_system is not None:
        _flush(current_system)

    print(f"  [loader] Energy sheet: {list(result.keys())} systems parsed")
    return result


# ──────────────────────────────────────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────────────────────────────────────

def load_data() -> dict:
    """
    Load and parse data.xlsx, returning BOM DataFrames plus energy data.

    Sections are located by scanning column B of Part 1 for known header
    strings.  Values are stored as-is from the Excel cells — no numeric
    coercion is applied.

    Returns
    -------
    dict with keys:
        "electrical"     – pd.DataFrame (equipment rows for the electrical system)
        "mechanical"     – pd.DataFrame (equipment rows for the mechanical system)
        "hybrid"         – pd.DataFrame (equipment rows for the hybrid system)
        "battery_lookup" – pd.DataFrame (battery fraction vs. tank fraction lookup)
        "tds_lookup"     – pd.DataFrame with columns ["tds_ppm", "ro_energy_kw"], 20 rows
        "depth_lookup"   – pd.DataFrame with columns ["depth_m", "pump_energy_kw"], 20 rows
        "energy"         – dict grouped by system ("mechanical", "electrical", "hybrid"),
                           each containing subsystems list, total_shaft_power,
                           total_turbine_input, and selected_turbine_kw.
                           May be None if the Energy sheet is absent from data.xlsx;
                           callers should fall back to SUBSYSTEM_POWER from config.py.

    Raises
    ------
    FileNotFoundError
        If data.xlsx does not exist at the configured path.
    ValueError
        If one or more expected section headers are missing from 'Part 1',
        or if parsing otherwise fails.
    """
    # ── 1. File existence check ──────────────────────────────────────────────
    if not DATA_FILE.exists():
        raise FileNotFoundError(
            f"data.xlsx not found at expected path: {DATA_FILE}\n"
            "Ensure the file is in the project root directory."
        )

    # ── 2. Open workbook ─────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)

    # ── 2b. Parse Part 2 lookup tables ──────────────────────────────────────
    tds_df, depth_df = _parse_part2_lookups(wb)

    if "Part 1" not in wb.sheetnames:
        raise ValueError(
            f"Expected sheet 'Part 1' not found. "
            f"Available sheets: {wb.sheetnames}"
        )
    ws = wb["Part 1"]

    # ── 3. Scan for section header rows ──────────────────────────────────────
    section_row_map: dict[str, int] = {}
    for row in ws.iter_rows(min_col=2, max_col=2):
        cell = row[0]
        if cell.value in SECTION_HEADERS:
            canonical_key = SECTION_HEADERS[cell.value]
            section_row_map[canonical_key] = cell.row
            print(f"  [loader] Found section '{cell.value}' at row {cell.row}")

    # ── 4. Validate all sections present ─────────────────────────────────────
    required = {"electrical", "mechanical", "hybrid"}
    missing = required - set(section_row_map.keys())
    if missing:
        raise ValueError(
            f"Missing section(s) in 'Part 1': {sorted(missing)}. "
            f"Found: {sorted(section_row_map.keys())}. "
            "Check that data.xlsx has not been modified."
        )

    # ── 5. Parse equipment sections ──────────────────────────────────────────
    elec_start   = section_row_map["electrical"]
    mech_start   = section_row_map["mechanical"]
    hybrid_start = section_row_map["hybrid"]

    electrical_rows = _parse_section(ws, elec_start,   stop_rows={mech_start},   cost_col=5)
    mechanical_rows = _parse_section(ws, mech_start,   stop_rows={hybrid_start}, cost_col=4)
    hybrid_rows     = _parse_section(ws, hybrid_start, stop_rows=set(),           cost_col=4)

    print(f"  [loader] Electrical: {len(electrical_rows)} equipment rows parsed")
    print(f"  [loader] Mechanical: {len(mechanical_rows)} equipment rows parsed")
    print(f"  [loader] Hybrid:     {len(hybrid_rows)} equipment rows parsed")

    # ── 6. Parse battery/tank lookup ─────────────────────────────────────────
    battery_df = _parse_battery_lookup(ws)
    print(f"  [loader] Battery lookup: {len(battery_df)} rows parsed")

    # ── 7. Parse Energy sheet ─────────────────────────────────────────────────
    energy_data = _parse_energy_sheet(wb)

    return {
        "electrical":    pd.DataFrame(electrical_rows, columns=EQUIPMENT_COLUMNS),
        "mechanical":    pd.DataFrame(mechanical_rows, columns=EQUIPMENT_COLUMNS),
        "hybrid":        pd.DataFrame(hybrid_rows,     columns=EQUIPMENT_COLUMNS),
        "battery_lookup": battery_df,
        "tds_lookup":    tds_df,
        "depth_lookup":  depth_df,
        "energy":        energy_data,
    }
